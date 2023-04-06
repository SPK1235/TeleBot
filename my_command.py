from datetime import datetime
import sqlite3
import openpyxl
import random
import requests
import asyncio
import aiohttp
from bs4 import BeautifulSoup
from PIL import Image, ImageDraw, ImageFont, ImageFilter


def command_start_time():
    """Получение поры дня при команде старт"""
    current_time = datetime.now().strftime("%H:%M")
    current_time = list(map(int, current_time.split(':')))
    current_time = current_time[0] + (current_time[1] / 60)
    time_day = None
    if 0 <= current_time <= 5:
        time_day = 'night'
    elif 5 < current_time <= 6.5:
        time_day = 'early_morning'
    elif 6.5 < current_time <= 10:
        time_day = 'early_morning'
    elif 10 < current_time <= 19:
        time_day = 'day'
    elif 19 < current_time <= 24:
        time_day = 'evening'
    return time_day


def creation_my_database():
    """Создание базы данных"""
    con = sqlite3.connect("my_database.db")
    cur = con.cursor()
    cur.execute("CREATE TABLE IF NOT EXISTS times_day (night, early_morning, morning, day, evening)")
    file_to_read = openpyxl.load_workbook('my_list.xlsx')
    sheet = file_to_read.active
    data = []
    rows = sheet.max_row
    cols = sheet.max_column
    for row in range(2, rows + 1):
        for col in range(1, cols + 1):
            value = sheet.cell(row, col).value
            data.append(value)
        cur.execute("INSERT INTO times_day VALUES (?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4]))
        data = []
    con.commit()
    con.close()


def read_database(time_day):
    """Чтение данных из базы данных"""
    con = sqlite3.connect("my_database.db")
    cur = con.cursor()
    result_column = cur.execute(f'SELECT {time_day} FROM times_day').fetchall()
    random_number = random.randrange(0, 10)
    return result_column[random_number][0]


def line_division(line: str):
    my_list = []
    start = 0
    for i in range(1, len(line)):
        if line[i].isalpha() and line[i].isupper():
            my_list.append(line[start:i])
            start = i
        elif i > start and line[i].isdigit() and int(line[i]) != 0 and len(my_list) != 0:
            my_list.append(line[start:i - 1])
            my_list.append(line[i - 1:])
        elif i > start and line[i].isdigit() and int(line[i]) == 0 and len(my_list) != 0:
            my_list.append(line[start:i])
            my_list.append(line[i:])
    return my_list[:4]


def line_division_2(line: str):
    my_list = line.split()
    for i in range(len(my_list)):
        if my_list[i] == 'Метров':
            my_list[i] = 'м'
        elif my_list[i] == 'в':
            my_list[i] = '/'
        elif my_list[i] == 'секунду,':
            my_list[i] = 'с , '
        elif my_list[i] == 'Ветер:':
            my_list[i] = 'Ветер: '
    my_str = ''.join(my_list)
    return my_str


def line_division_3(line: str):
    my_list = line.split()
    for i in range(len(my_list)):
        if my_list[i] == 'Миллиметров':
            my_list[i] = ' мм '
        elif my_list[i] == 'ртутного':
            my_list[i] = 'рт.'
        elif my_list[i] == 'столба.':
            my_list[i] = 'ст.'
        elif my_list[i] == 'Давление:':
            my_list[i] = 'Давление: '
    my_str = ''.join(my_list)
    return my_str


def line_division_4(line: str):
    line = line.split(',')
    del line[0]
    del line[0]
    for i in range(len(line[1])):
        if line[1][i].isdigit():
            if line[1][i - 1] == ' ':
                line[1] = line[1][:i] + '+' + line[1][i:]
                break
    for i in range(len(line[2])):
        if line[2][i].isdigit():
            if line[2][i - 1] == ' ':
                line[2] = line[2][:i] + '+' + line[2][i:]
                break
    line[1] = line[1] + 'C'
    line[2] = line[2] + 'C'
    return line


def parser_1(page):
    global flag_moscow
    query_result = {}
    try:
        location = page.find('h1', attrs={'class': 'title title_level_1 header-title__title'}).text
        print(location)
        location = list(location.split(','))
        if len(location) == 2:
            query_result['city'] = location[1]
            if location[1].strip() == 'Москва':
                flag_moscow = True
            query_result['area'] = location[0]
        elif len(location) == 1:
            query_result['city'] = location[0]
            query_result['area'] = ''
        now_time_div = page.find('div', attrs={'class': 'fact__time-yesterday-wrap'})
        now_time = now_time_div.find('time', attrs={'class': 'time fact__time'}).text
        query_result['time'] = now_time
        degree = now_time_div.find('span', attrs={'class': 'a11y-hidden'}).text
        degree = degree[-1] + 'C'
        temperature_yesterday = (now_time_div.find('div', attrs={'class': 'term__label'}).text + ' ' +
                                 now_time_div.find('span', attrs={'class': 'temp__value temp__value_with-unit'}).text +
                                 degree)
        query_result['temperature_yesterday'] = temperature_yesterday
        weather_now = page.find('div', attrs={'class': 'fact__temp-wrap'}).text
        weather_now = line_division(weather_now)
        weather_now[0] = weather_now[0] + degree
        weather_now[-1] = weather_now[-1] + degree
        query_result['now'] = weather_now
        wind_humidity_pressure = page.find('div', attrs={'class': 'fact__props'})
        wind = wind_humidity_pressure.find('div', attrs={'class': 'term term_orient_v fact__wind-speed'})
        wind = wind.find('span', attrs={'class': 'a11y-hidden'}).text
        wind = line_division_2(wind)
        query_result['wind'] = wind
        humidity = wind_humidity_pressure.find('div', attrs={'class': 'term term_orient_v fact__humidity'})
        humidity = humidity.find('span', attrs={'class': 'a11y-hidden'}).text
        query_result['humidity'] = humidity
        pressure = wind_humidity_pressure.find('div', attrs={'class': 'term term_orient_v fact__pressure'})
        pressure = pressure.find('span', attrs={'class': 'a11y-hidden'}).text
        pressure = line_division_3(pressure)
        query_result['pressure'] = pressure
        weather = page.find('div', attrs={'class': 'content content_compressed i-bem'})
        weather = weather.find('div', attrs={'class': 'forecast-briefly__days swiper-container'})
        weather = weather.find('ul', attrs={'class': 'swiper-wrapper'})
        weather = weather.find_all('li')
        weather_today = weather[1].find('a').get('aria-label')
        weather_today = line_division_4(weather_today)
        weather_tomorrow = weather[2].find('a').get('aria-label')
        weather_tomorrow = line_division_4(weather_tomorrow)
        query_result['weather_today'] = weather_today
        query_result['weather_tomorrow'] = weather_tomorrow
        query_result['flag_status'] = True
    except:
        query_result['flag_status'] = False
    return query_result


def parser_yandex_weather(latitude, longitude):
    query_result = {}
    global flag_moscow
    url = f'https://yandex.ru/pogoda/?lat={latitude}&lon={longitude}'
    header = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
              'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:106.0) Gecko/20100101 Firefox/106.0'}
    resp = requests.get(url, headers=header)
    try:
        if resp.status_code == 200:
            page = BeautifulSoup(resp.content, "html.parser")
            try:
                query_result = parser_1(page)
            except:
                if flag_moscow:
                    latitude = '55.755863'
                    longitude = '37.6177'
                    url = f'https://yandex.ru/pogoda/?lat={latitude}&lon={longitude}'
                    resp = requests.get(url, headers=header)
                    page = BeautifulSoup(resp.content, "html.parser")
                    query_result = parser_1(page)
            flag_moscow = False
            return query_result
    except:
        query_result['flag_status'] = False
        flag_moscow = False
        return query_result


def color_selection(random_number):
    color_list = ['Yellow', 'Aqua', 'Lime']
    if random_number in [1, 2, 3, 4, 5, 8, 11, 12]:
        return color_list[0]
    elif random_number in [6, 7, 9, 10]:
        return color_list[1]


def creating_picture(query_result):
    width = 1080
    random_number = random.randrange(1, 13)
    color = color_selection(random_number)
    picture = Image.open(f"picture/{random_number}.jpg").resize((1080, 1980))
    picture = picture.filter(ImageFilter.BLUR)
    inscription = ImageDraw.Draw(picture)
    if query_result['flag_status']:
        geo_picture = Image.open("picture/geo.png")
        picture.paste(geo_picture, (130, 100), mask=geo_picture)
        text = query_result['city']
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((200, 110), text, fill=color, font=font)
        text = query_result['area']
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((215, 185), text, fill=color, font=font)
        text = query_result['time']
        text = text[:-2]
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 310)
        text = query_result['temperature_yesterday']
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 440)
        text = 'Погода за окном :'
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 550)
        text = query_result['now'][0] # в одну строку 1-й элемент
        font = ImageFont.truetype('arial.ttf', size=100)
        inscription.text((100, 670), text, fill=color, font=font)
        text = query_result['now'][1] # в одну строку 2-й элемент
        font = ImageFont.truetype('arial.ttf', size=55)
        size_text = inscription.textlength(text, font=font)
        x = (width + 335) // 2 - (size_text // 2)
        inscription.text((x, 700), text, fill=color, font=font)
        text = query_result['now'][2] + ' :'
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((130, 822), text, fill=color, font=font)
        text = query_result['now'][3]
        font = ImageFont.truetype('arial.ttf', size=80)
        inscription.text((660, 815), text, fill=color, font=font)
        text = query_result['wind']
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((130, 940), text, fill=color, font=font)
        text = query_result['humidity']
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((130, 1040), text, fill=color, font=font)
        text = query_result['pressure']
        font = ImageFont.truetype('arial.ttf', size=55)
        inscription.text((130, 1140), text, fill=color, font=font)
        text = 'Сегодня :'
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1240)
        text = query_result['weather_today'][0]
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1340)
        text = query_result['weather_today'][1] + ' ;  ' + query_result['weather_today'][2]
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1440)
        text = 'Завтра :'
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1540)
        text = query_result['weather_tomorrow'][0]
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1640)
        text = query_result['weather_tomorrow'][1] + ' ;  ' + query_result['weather_tomorrow'][2]
        table(picture, text, 'arial.ttf', 55, color, 0, 1080, 1740)
        text = 'Погода на сайте "Yandex.ru"'
        table(picture, text, 'beer-money12.ttf', 55, color, 0, 1080, 1880)
        #picture.show()
        picture.save('1.jpg')
    else:
        text = 'Сайт "Yandex.ru" недоступен'
        table(picture, text, 'beer-money12.ttf', 80, color, 0, 1080, 500)
        text = 'ERROR'
        table(picture, text, 'arial.ttf', 120, color, 0, 1080, 800)
        #picture.show()
        picture.save('1.jpg')


def parser_myfin_by_minsk():
    query_result = {}
    url = 'https://myfin.by/currency/minsk'
    header = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
              'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:106.0) Gecko/20100101 Firefox/106.0'}
    resp = requests.get(url, headers=header)
    try:
        if resp.status_code == 200:
            query_result['flag_status'] = True
            page = BeautifulSoup(resp.content, "html.parser")
            exchange_rates_block = page.find('div', attrs={'class': 'g-wrap outer-bg'})
            data_now = page.find('div', attrs={'class': 'c-section-wrapper'})
            data_now = data_now.find('h1').text
            data_now = list(data_now.split())
            data_now = ' '.join(data_now[:2] + data_now[4:])
            query_result['data_now'] = data_now
            exchange_rates_block = exchange_rates_block.find('div', attrs={'class': 'c-best-rates'})
            exchange_rates = exchange_rates_block.find_all('tr')
            for currency in exchange_rates[1:4]:
                currency_name = currency.find('a').get_text()
                currency_name = list(currency_name.split())
                currency_name = ' '.join(currency_name[:2])
                query_result[currency_name] = []
                volume = currency.find_all('td')[1:]
                for i in volume:
                    query_result[currency_name].append(i.text)
        return query_result
    except:
        query_result['flag_status'] = False
        return query_result


def parser_myfin_by_moscow():
    query_result = {}
    url = 'https://myfin.by/currency/moskva'
    header = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
              'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:106.0) Gecko/20100101 Firefox/106.0'}
    resp = requests.get(url, headers=header)
    try:
        if resp.status_code == 200:
            query_result['flag_status'] = True
            page = BeautifulSoup(resp.content, "html.parser")
            exchange_rates_block = page.find('div', attrs={'class': 'g-wrap outer-bg'})
            exchange_rates_block = exchange_rates_block.find('div', attrs={'class': 'c-container'})
            exchange_rates_block = exchange_rates_block.find('div', attrs={'class': 'c-section bg-pearl-white'})
            exchange_rates = exchange_rates_block.find_all('tr')
            for currency in exchange_rates[1:]:
                currency_name = currency.find('a').get_text()
                query_result[currency_name] = []
                volume = currency.find_all('td')[1:-1]
                for i in volume:
                    query_result[currency_name].append(i.text)
        return query_result
    except:
        query_result['flag_status'] = False
        return query_result


def table(picture, text, font_ttf, size, color, x_1, x_2, y):
    inscription = ImageDraw.Draw(picture)
    font = ImageFont.truetype(font_ttf, size=size)
    size_text = inscription.textlength(text, font=font)
    x = (x_2 - x_1 - size_text) // 2 + x_1
    inscription.text((x, y), text, fill=color, font=font)


def creating_picture_currency():
    query_result = parser_myfin_by_minsk()
    query_result_2 = parser_myfin_by_moscow()
    random_number = random.randrange(1, 13)
    color = color_selection(random_number)
    picture = Image.open(f"picture/{random_number}.jpg").resize((1080, 1980))
    picture = picture.filter(ImageFilter.BLUR)
    inscription = ImageDraw.Draw(picture)
    if query_result['flag_status'] and query_result_2['flag_status']:
        text = query_result['data_now']
        table(picture, text, 'arial.ttf', 60, color, 0, 1080, 150)
        text = 'Минск :'
        table(picture, text, 'arial.ttf', 65, color, 0, 1080, 300)
        text = 'Валюта'
        table(picture, text, 'arial.ttf', 35, color, 0, 360, 430)
        text = 'Покупка'
        table(picture, text, 'arial.ttf', 35, color, 360, 560, 430)
        text = 'Продажа'
        table(picture, text, 'arial.ttf', 35, color, 560, 760, 430)
        text = 'НБ РБ'
        table(picture, text, 'arial.ttf', 35, color, 760, 960, 430)
        trend_picture = Image.open("picture/red.png").resize((25, 25)).convert("RGBA")
        picture.paste(trend_picture, (1000, 452), mask=trend_picture)
        trend_picture = Image.open("picture/green.png").resize((25, 25)).convert("RGBA")
        picture.paste(trend_picture, (1000, 418), mask=trend_picture)
        line = ImageDraw.Draw(picture)
        line.line((20, 490, 1060, 490), fill=color, width=3)
        line.line((360, 430, 360, 810), fill=color, width=3)  # 1я линия
        line.line((560, 430, 560, 810), fill=color, width=3)  # 2я линия
        line.line((760, 430, 760, 810), fill=color, width=3)  # 3я линия
        line.line((960, 430, 960, 810), fill=color, width=3)  # 4я линия
        line.line((20, 600, 1060, 600), fill=color, width=3)
        line.line((20, 710, 1060, 710), fill=color, width=3)
        text = 'Доллар США'
        font = ImageFont.truetype('arial.ttf', size=45)
        inscription.text((30, 520), text, fill=color, font=font)
        text = 'Евро'
        font = ImageFont.truetype('arial.ttf', size=45)
        inscription.text((30, 630), text, fill=color, font=font)
        text = 'Рос. рубль'
        font = ImageFont.truetype('arial.ttf', size=45)
        inscription.text((30, 740), text, fill=color, font=font)
        text = query_result['Доллар США'][0]
        table(picture, text, 'arial.ttf', 45, color, 360, 560, 520)
        text = query_result['Доллар США'][1]
        table(picture, text, 'arial.ttf', 45, color, 560, 760, 520)
        text = query_result['Доллар США'][2]
        table(picture, text, 'arial.ttf', 45, color, 760, 960, 520)
        text = query_result['Евро'][0]
        table(picture, text, 'arial.ttf', 45, color, 360, 560, 630)
        text = query_result['Евро'][1]
        table(picture, text, 'arial.ttf', 45, color, 560, 760, 630)
        text = query_result['Евро'][2]
        table(picture, text, 'arial.ttf', 45, color, 760, 960, 630)
        text = query_result['Российский рубль'][0]
        table(picture, text, 'arial.ttf', 45, color, 360, 560, 740)
        text = query_result['Российский рубль'][1]
        table(picture, text, 'arial.ttf', 45, color, 560, 760, 740)
        text = query_result['Российский рубль'][2]
        table(picture, text, 'arial.ttf', 45, color, 760, 960, 740)
        trend = float(list(query_result['Доллар США'][3].split())[1])
        if 0 <= trend:
            trend_picture = Image.open("picture/green.png").resize((50, 50)).convert("RGBA")
        else:
            trend_picture = Image.open("picture/red.png").resize((50, 50)).convert("RGBA")
        picture.paste(trend_picture, (990, 520), mask=trend_picture)
        trend = float(list(query_result['Евро'][3].split())[1])
        if 0 <= trend:
            trend_picture = Image.open("picture/green.png").resize((50, 50)).convert("RGBA")
        else:
            trend_picture = Image.open("picture/red.png").resize((50, 50)).convert("RGBA")
        picture.paste(trend_picture, (990, 630), mask=trend_picture)
        trend = float(list(query_result['Российский рубль'][3].split())[1])
        if 0 <= trend:
            trend_picture = Image.open("picture/green.png").resize((50, 50)).convert("RGBA")
        else:
            trend_picture = Image.open("picture/red.png").resize((50, 50)).convert("RGBA")
        picture.paste(trend_picture, (990, 740), mask=trend_picture)
        text = 'Москва :'
        table(picture, text, 'arial.ttf', 65, color, 0, 1080, 950)
        text = 'Валюта'
        table(picture, text, 'arial.ttf', 35, color, 0, 360, 1100)
        text = 'Покупка'
        table(picture, text, 'arial.ttf', 35, color, 360, 560, 1100)
        text = 'Продажа'
        table(picture, text, 'arial.ttf', 35, color, 560, 760, 1100)
        text = 'ЦБ РФ'
        table(picture, text, 'arial.ttf', 35, color, 760, 960, 1100)
        line.line((20, 1150, 1060, 1150), fill=color, width=3)  # 1я линия гор.
        line.line((20, 1260, 1060, 1260), fill=color, width=3)  # 2я линия гор.
        line.line((360, 1100, 360, 1360), fill=color, width=3)  # 1я линия верт.
        line.line((560, 1100, 560, 1360), fill=color, width=3)  # 2я линия верт.
        line.line((760, 1100, 760, 1360), fill=color, width=3)  # 3я линия верт.
        line.line((960, 1100, 960, 1360), fill=color, width=3)  # 4я линия верт.
        text = 'Доллар США'
        font = ImageFont.truetype('arial.ttf', size=45)
        inscription.text((30, 1180), text, fill=color, font=font)
        text = 'Евро'
        font = ImageFont.truetype('arial.ttf', size=45)
        inscription.text((30, 1290), text, fill=color, font=font)
        text = query_result_2['Доллар США'][1]
        table(picture, text, 'arial.ttf', 45, color, 360, 560, 1180)
        text = query_result_2['Доллар США'][2]
        table(picture, text, 'arial.ttf', 45, color, 560, 760, 1180)
        text = query_result_2['Доллар США'][0]
        table(picture, text, 'arial.ttf', 45, color, 760, 960, 1180)
        text = query_result_2['Евро'][1]
        table(picture, text, 'arial.ttf', 45, color, 360, 560, 1290)
        text = query_result_2['Евро'][2]
        table(picture, text, 'arial.ttf', 45, color, 560, 760, 1290)
        text = query_result_2['Евро'][0]
        table(picture, text, 'arial.ttf', 45, color, 760, 960, 1290)
        text = 'Курс валют на сайте "Myfin.by"'
        table(picture, text, 'beer-money12.ttf', 55, color, 0, 1080, 1880)
        owl_picture = Image.open("picture/owl.png").resize((400, 400)).convert("RGBA")
        picture.paste(owl_picture, (350, 1450), mask=owl_picture)
        #picture.show()
        picture.save('1.jpg')
    else:
        pass


flag_moscow = False


